from tkinter import * 
from openpyxl.workbook import Workbook
from openpyxl import load_workbook


root = Tk()
root.title('Tkinter GUI - Sailendra')
root.geometry('400x400')

# Create workbook instance
wb = Workbook()

#Load exiting workbook
wb = load_workbook('Samples/customers.csv') 

# Create active worksheet
ws = wb.active

# Create variable for  column A
column_a = ws['A']
column_b = ws['B']


root.mainloop()